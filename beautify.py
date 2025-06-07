import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np

#已合并至tools.py

def beautify_excel(input_file, output_file, header=True, thousands_sep=True, auto_fit=True):
    """
    美化Excel文件的函数，主要用于美化mc_cal后返回的mc.xlsx
    标题行、文本列居中，数据列右对齐，加边框，数据加千分位分割，自动调整列宽
    
    参数:
    1.input_file (str): 输入Excel文件路径
    2.output_file (str): 输出Excel文件路径
    3.header (bool): 是否包含标题行
    4.thousands_sep (bool): 是否添加千位分隔符
    5.auto_fit (bool): 是否自动调整列宽
    """

    # 加载工作簿
    df = pd.read_excel(input_file)
    wb = load_workbook(input_file)
    ws = wb.active
    
    # 设置字体
    font = Font(name='Calibri', size=11)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
    
    # 确定数据类型列（假设标题行在第1行）
    first_data_row = 2 if header else 1
    text_columns = set()  # 存储文本列索引
    numeric_columns = set()  # 存储数字列索引
    
    # 检查每一列的数据类型（基于前20行样本）
    sample_size = min(20, len(df))
    for col_idx, col in enumerate(df.columns):
        is_text = True
        for row_idx in range(sample_size):
            if row_idx >= len(df):
                continue
            value = df.iloc[row_idx, col_idx]
            if isinstance(value, (int, float)) and not np.isnan(value):
                is_text = False
                break
        if is_text:
            text_columns.add(col_idx)
        else:
            numeric_columns.add(col_idx)
    
    # 设置对齐方式
    title_alignment = Alignment(horizontal='center', vertical='center')  # 标题行居中
    text_alignment = Alignment(horizontal='center', vertical='center')    # 文本列居中
    numeric_alignment = Alignment(horizontal='right', vertical='center')  # 数字列右对齐
    
    # 标题行对齐
    if header:
        for cell in ws[1]:
            cell.alignment = title_alignment
    
    # 数据行对齐（根据列类型）
    for row in ws.iter_rows(min_row=first_data_row):
        for cell in row:
            col_idx = cell.column - 1  # 转换为0-based索引
            if col_idx in text_columns:
                cell.alignment = text_alignment
            elif col_idx in numeric_columns:
                cell.alignment = numeric_alignment
    
    # 设置边框
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    
    # 添加千位分隔符（仅对数字列）
    if thousands_sep:
        for row in ws.iter_rows(min_row=first_data_row):
            for cell in row:
                col_idx = cell.column - 1
                if col_idx in numeric_columns and isinstance(cell.value, (int, float)):
                    if isinstance(cell.value, int):
                        cell.number_format = '#,##0'  # 整数格式
                    else:
                        cell.number_format = '#,##0.00'  # 浮点数格式
    
    # 自动调整列宽
    if auto_fit:
        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    if cell.value:
                        cell_width = len(str(cell.value))
                        if cell_width > column_widths[i]:
                            column_widths[i] = cell_width
                except IndexError:
                    column_widths.append(len(str(cell.value)))
        
        # 设置列宽
        for i, width in enumerate(column_widths):
            column_letter = get_column_letter(i + 1)
            adjusted_width = min(max(width * 1.2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存美化后的文件
    wb.save(output_file)
    print(f"优化后Excel文件已并保存至: {output_file}")

# 使用示例
if __name__ == "__main__":
    beautify_excel(
        input_file='20250430mc.xlsx',
        output_file='beautified_file.xlsx',
        header=True,
        thousands_sep=True,
        auto_fit=True
    )